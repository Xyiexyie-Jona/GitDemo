import openpyxl
class HomePageData:

    test_HomePage_Data = [{"firstname":"Jona", "email": "j@gmail.com", "password": "123", "gender": "Female"},
                            {"firstname": "Rahul", "email": "shetty@gmail.com", "password": "456", "gender": "Male"}]


    #@staticmethod does not require self parameter
    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\asus tuf\\PycharmProjects\\PythonDemo.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(1, sheet.max_column + 1):  # to get columns
                    # print(sheet.cell(row= i, column=j).value)
                    # Dict["firstname"] = "Jona"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

                print(Dict)
        return [Dict]
